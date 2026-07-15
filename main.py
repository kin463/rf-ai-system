import os
import re
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook

app = FastAPI()
app.mount("/static", StaticFiles(directory="."), name="static")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

class ChatRequest(BaseModel):
    message: str
    mode: str

# Excelを読み込む関数
def read_excel_all():
    wb = load_workbook("master.xlsx")
    # シート2：帰社日検索マスタ
    ws_kisha = wb["2. 帰社日検索マスタ"]
    kisha_data = []
    for row in ws_kisha.iter_rows(min_row=2, values_only=True):
        fullname, dept, kisha_date = row
        if fullname is None:
            continue
        kisha_data.append([fullname.strip(), dept if dept else "", kisha_date if kisha_date is not None else ""])

    # シート3：勤怠規定マスタ（A規定ID B大分類 C小分類 D規定内容 E適用条件 F備考）
    ws_rule = wb["3. 勤怠規定マスタ"]
    rule_data = []
    for row in ws_rule.iter_rows(min_row=2, values_only=True):
        rule_id, big_class, small_class, content, condition, note = row
        if small_class is None or content is None:
            continue
        rule_data.append([small_class.strip(), content.strip()])

    # シート4：結婚祝金・有給休暇計算マスタ
    ws_calc = wb["4. 結婚祝金･有給休暇計算マスタ"]
    calc_data = []
    for row in ws_calc.iter_rows(min_row=2, values_only=True):
        kind, years, result = row
        calc_data.append([kind, years, result])
    wb.close()
    return kisha_data, rule_data, calc_data

# 文章から数字抽出
def extract_year(text: str):
    nums = re.findall(r"\d+", text)
    if nums:
        return int(nums[0])
    return None

# 質問文から社員名抽出「XXの」「XXが」に対応
def extract_name(input_text: str):
    match = re.search(r"([一-龥]{2,6})(の|が)", input_text)
    if match:
        return match.group(1).strip()
    return input_text.strip()

@app.get("/health")
async def health():
    return {"status": "success", "message": "RF AI System is online."}

@app.get("/")
async def root():
    html_file = os.path.join(os.path.dirname(__file__), "index.html")
    return FileResponse(html_file)

@app.post("/api/chat")
async def chat(request: ChatRequest):
    try:
        input_msg = request.message.strip()
        kisha_data, rule_data, calc_data = read_excel_all()

        # ========== kishaモード：帰社日検索 ==========
        if request.mode == "kisha":
            search_name = extract_name(input_msg)
            find_result = []
            for fullname, dept, kisha_date in kisha_data:
                if search_name in fullname:
                    find_result.append([fullname, dept, kisha_date])
            if len(find_result) == 0:
                return {"response": "該当するメンバーが見つかりませんでした。"}
            output_lines = []
            for name, dept, dt in find_result:
                if dt == "":
                    output_lines.append(f"{name} {dept}：帰社日は設定されていません")
                else:
                    output_lines.append(f"{name} {dept}：帰社日：{dt}")
            reply_text = "ご確認いただきありがとうございます。該当者の帰社日は以下です。\n" + "\n".join(output_lines)
            return {"response": reply_text}

        # ========== faqモード：勤怠規定検索 ==========
        else:
            question = input_msg
            years = extract_year(question)

            # 1.結婚祝金の判定
            if "結婚" in question:
                if years is not None:
                    if years < 2:
                        return {"response": "2万円"}
                    elif 2 <= years < 5:
                        return {"response": "3万円"}
                    else:
                        return {"response": "5万円"}
                else:
                    return {"response": "勤続年数を数字で記入してください。"}

            # 2.有給休暇判定
            if "有給休暇" in question or "年次有給" in question:
                if years is not None:
                    if years == 0:
                        return {"response": "10日"}
                    elif years == 1:
                        return {"response": "11日"}
                    elif years == 2:
                        return {"response": "12日"}
                    elif years == 3:
                        return {"response": "14日"}
                    elif years == 4:
                        return {"response": "16日"}
                    elif years == 5:
                        return {"response": "18日"}
                    elif years >= 6:
                        return {"response": "20日"}

            # 3.Excelの小分類から回答を取得
            answer = ""
            for small_class, rule_content in rule_data:
                if small_class in question:
                    answer = rule_content
                    break
            if answer != "":
                return {"response": answer}
            else:
                return {"response": "資料に記載がありません"}

    except Exception as e:
        print("エラー詳細：", str(e))
        return {"response": f"サーバーエラー：{str(e)}"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
